#!/usr/bin/env python3
"""
Phase 7: Create Professional XLSX Spreadsheet
Color-coded, styled, sorted by composite score
"""

import json
import sys
import subprocess
import os

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# Color schemes
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
IMMEDIATE_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")  # Green
HIGH_FILL = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")       # Light green
MODERATE_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")   # Yellow
WATCHLIST_FILL = None  # No fill

BOLD_FONT = Font(bold=True)
LINK_FONT = Font(color="0563C1", underline="single")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF")
)


def fmt_price(val):
    if not val:
        return "Not stated"
    try:
        return f"${int(val):,}"
    except:
        return str(val)


def fmt_int(val):
    if val is None or val == "Not stated":
        return "Not stated"
    try:
        v = int(float(val))
        return f"{v:,}" if v > 0 else "Not stated"
    except:
        return str(val) if val else "Not stated"


def fmt_bool(val):
    if val is True:
        return "Yes"
    # Treat False, None, 0, and missing values as "No"
    return "No"


def get_block_context(prop, clusters):
    """Build block context string."""
    parts = []
    if prop.get("_in_cluster"):
        size = prop.get("_cluster_size", "")
        combined = prop.get("_cluster_combined_ask", "")
        if size:
            parts.append(f"{size} qualifying lots on block")
        if combined and isinstance(combined, (int, float)):
            parts.append(f"Combined ask: ${combined:,.0f}")
    if prop.get("_block_llc_deed"):
        parts.append("LLC deed transfer on block (last 12mo)")
    if prop.get("_block_demo"):
        parts.append("Demo permit on block (last 6mo)")
    if prop.get("_block_new_building"):
        parts.append("New building permit on block (last 6mo)")
    return "; ".join(parts) if parts else "None"


def get_notes(prop):
    """Build notes string."""
    notes = []
    reasons = prop.get("_score_reasons", [])
    for r in reasons:
        if r not in ["Active for-sale listing (+3)"]:  # Skip obvious ones
            notes.append(r)
    if prop.get("_data_quality_note"):
        notes.append(f"⚠ {prop['_data_quality_note']}")
    return "; ".join(notes) if notes else "Active listing in R7+ zone"


def main():
    with open("/tmp/nyc_assemblage/verified_properties.json") as f:
        properties = json.load(f)

    with open("/tmp/nyc_assemblage/clusters.json") as f:
        clusters = json.load(f)

    print(f"Creating XLSX for {len(properties)} properties...", file=sys.stderr)

    wb = openpyxl.Workbook()

    # ============================================================
    # Sheet 1: All Properties
    # ============================================================
    ws = wb.active
    ws.title = "All Properties"

    headers = [
        "Borough", "Property Address", "Asking Price", "Units (Res)",
        "Lot Size (SF)", "Building SF", "Year Built", "Zoning",
        "Composite Score", "Priority",
        "Tax Lien?", "Lis Pendens?", "HPD Violations",
        "Demo on Block?", "NB on Block?", "LLC Deed on Block?",
        "Estate Signal?", "Federal Lien?",
        "311 Complaints", "ECB Violations", "FDNY Vacate?",
        "DOB Complaints", "Block CO?", "CitiBike Stations",
        "Price Drop %", "Listing Cycles", "Days on Market",
        "Block Context", "Zillow Link", "ZoLa Link", "Notes"
    ]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Set column widths
    col_widths = {
        1: 12,   # Borough
        2: 40,   # Address
        3: 15,   # Price
        4: 10,   # Units
        5: 12,   # Lot Size
        6: 12,   # Bldg SF
        7: 10,   # Year Built
        8: 10,   # Zoning
        9: 12,   # Score
        10: 12,  # Priority
        11: 10,  # Tax Lien
        12: 12,  # Lis Pendens
        13: 14,  # HPD
        14: 14,  # Demo
        15: 12,  # NB
        16: 16,  # LLC Deed
        17: 14,  # Estate
        18: 13,  # Federal Lien
        19: 14,  # 311 Complaints
        20: 14,  # ECB Violations
        21: 12,  # FDNY Vacate
        22: 14,  # DOB Complaints
        23: 10,  # Block CO
        24: 14,  # CitiBike Stations
        25: 12,  # Price Drop
        26: 14,  # Cycles
        27: 14,  # DOM
        28: 35,  # Block Context
        29: 15,  # Zillow
        30: 15,  # ZoLa
        31: 50,  # Notes
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    # Write data rows
    for row_idx, prop in enumerate(properties, 2):
        priority = prop.get("_priority", "Watchlist")

        # Determine row fill
        if priority == "Immediate":
            row_fill = IMMEDIATE_FILL
        elif priority == "High":
            row_fill = HIGH_FILL
        elif priority == "Moderate":
            row_fill = MODERATE_FILL
        else:
            row_fill = None

        dom = prop.get("daysOnMarket", 0) or 0

        # Price drop
        se_drop = prop.get("se_price_drop_pct", 0) or 0
        price_drop_str = f"{se_drop:.0f}%" if se_drop > 0 else "N/A"

        # Listing cycles
        cycles = prop.get("se_cycles", 0) or 0
        cycles_str = str(cycles) if cycles > 0 else "N/A"

        row_data = [
            prop.get("_borough", "Not stated"),
            prop.get("address", "Not stated"),
            fmt_price(prop.get("price")),
            fmt_int(prop.get("_units_res")),
            fmt_int(prop.get("_lot_area")),
            fmt_int(prop.get("_bldg_area")),
            fmt_int(prop.get("_year_built")),
            prop.get("_zoning", "Not stated"),
            prop.get("_score", 0),
            priority,
            fmt_bool(prop.get("_tax_lien")),
            fmt_bool(prop.get("_acris_lis_pendens")),
            fmt_int(prop.get("_hpd_violations")),
            fmt_bool(prop.get("_block_demo")),
            fmt_bool(prop.get("_block_new_building")),
            fmt_bool(prop.get("_block_llc_deed")),
            fmt_bool(prop.get("_acris_estate")),
            fmt_bool(prop.get("_acris_federal_lien")),
            fmt_int(prop.get("_311_complaints")),
            fmt_int(prop.get("_ecb_violations")),
            fmt_bool(prop.get("_fdny_vacate")),
            fmt_int(prop.get("_dob_complaints")),
            fmt_bool(prop.get("_block_co")),
            fmt_int(prop.get("_citibike_stations")),
            price_drop_str,
            cycles_str,
            str(dom) if dom else "Not stated",
            get_block_context(prop, clusters),
            prop.get("zillowUrl", "Not stated"),
            prop.get("_zola_url", "Not stated"),
            get_notes(prop),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = TOP_ALIGN

            # Apply row color
            if row_fill:
                cell.fill = row_fill

            # Hyperlinks for URL columns
            if col_idx == 29 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.value = "Zillow"
                cell.font = LINK_FONT

            elif col_idx == 30 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.value = "ZoLa"
                cell.font = LINK_FONT

            # Score column: bold
            elif col_idx == 9:
                cell.font = BOLD_FONT
                cell.alignment = CENTER_ALIGN

            # Yes/No columns: center
            elif col_idx in [11, 12, 14, 15, 16, 17, 18, 21, 23]:
                cell.alignment = CENTER_ALIGN
                if value == "Yes":
                    cell.font = Font(bold=True, color="C00000")

        ws.row_dimensions[row_idx].height = 20

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ============================================================
    # Sheet 2: Immediate Priority
    # ============================================================
    ws2 = wb.create_sheet("Immediate Priority")
    immediate_props = [p for p in properties if p.get("_priority") == "Immediate"]

    # Title
    ws2.cell(row=1, column=1, value="IMMEDIATE PRIORITY PROPERTIES — NYC Assemblage Intelligence").font = Font(
        bold=True, size=14, color="1F4E79"
    )
    ws2.cell(row=2, column=1, value=f"Score 20+ | {len(immediate_props)} properties | Generated 2026-03-24")
    ws2.merge_cells("A1:H1")
    ws2.merge_cells("A2:H2")

    # Headers
    imm_headers = ["Address", "Borough", "Score", "Priority", "Price", "Lot SF",
                   "Zoning", "Yr Built", "Tax Lien", "Lis Pendens", "HPD Viol",
                   "Estate", "Block Signal", "Key Signals", "Zillow", "ZoLa"]
    for col_idx, h in enumerate(imm_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, prop in enumerate(immediate_props, 4):
        reasons_short = "; ".join(
            r for r in prop.get("_score_reasons", [])
            if "Active for-sale" not in r
        )[:120]
        row_data = [
            prop.get("address", ""),
            prop.get("_borough", ""),
            prop.get("_score", 0),
            prop.get("_priority", ""),
            fmt_price(prop.get("price")),
            fmt_int(prop.get("_lot_area")),
            prop.get("_zoning", ""),
            fmt_int(prop.get("_year_built")),
            fmt_bool(prop.get("_tax_lien")),
            fmt_bool(prop.get("_acris_lis_pendens")),
            fmt_int(prop.get("_hpd_violations")),
            fmt_bool(prop.get("_acris_estate")),
            get_block_context(prop, clusters)[:80],
            reasons_short,
            prop.get("zillowUrl", ""),
            prop.get("_zola_url", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = IMMEDIATE_FILL
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
            if col_idx in [15, 16] and value and value.startswith("http"):
                cell.hyperlink = value
                cell.value = "Link"
                cell.font = LINK_FONT

    # Column widths for sheet 2
    for col, width in enumerate([45, 12, 8, 10, 14, 10, 10, 8, 8, 10, 8, 8, 35, 60, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    # ============================================================
    # Sheet 3: Clusters
    # ============================================================
    ws3 = wb.create_sheet("Cluster Opportunities")
    ws3.cell(row=1, column=1, value="BLOCK CLUSTER OPPORTUNITIES — 2+ Qualifying Properties on Same Block")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E79")
    ws3.merge_cells("A1:G1")

    cluster_headers = ["Borough", "Block", "# Properties", "Addresses",
                       "Combined Ask", "Total Lot SF", "Zones", "Max Score", "Avg Score"]
    for col_idx, h in enumerate(cluster_headers, 1):
        cell = ws3.cell(row=2, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    with open("/tmp/nyc_assemblage/clusters.json") as f:
        clusters_data = json.load(f)

    for row_idx, cluster in enumerate(clusters_data, 3):
        combined_ask = cluster.get("combined_asking_price", "")
        combined_ask_str = fmt_price(combined_ask) if isinstance(combined_ask, (int, float)) else "Unable to verify"
        total_lot = cluster.get("total_lot_area_sf", "")
        total_lot_str = fmt_int(total_lot) if isinstance(total_lot, (int, float)) else "Unable to verify"

        row_data = [
            cluster.get("borough", ""),
            cluster.get("block", ""),
            cluster.get("property_count", 0),
            "\n".join(cluster.get("addresses", [])),
            combined_ask_str,
            total_lot_str,
            ", ".join(cluster.get("zones", [])),
            cluster.get("max_score", 0),
            cluster.get("avg_score", 0),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN

    for col, width in enumerate([12, 12, 10, 60, 16, 14, 20, 10, 10], 1):
        ws3.column_dimensions[get_column_letter(col)].width = width

    # ============================================================
    # Sheet 4: Signal Summary
    # ============================================================
    ws4 = wb.create_sheet("Signal Summary")
    ws4.cell(row=1, column=1, value="PRE-MARKET SIGNAL SUMMARY").font = Font(bold=True, size=14, color="1F4E79")
    ws4.merge_cells("A1:C1")

    signal_data = [
        ("Total R7+ Properties Analyzed", len(properties)),
        ("Immediate Priority (Score 20+)", sum(1 for p in properties if p.get("_priority") == "Immediate")),
        ("High Priority (Score 15-19)", sum(1 for p in properties if p.get("_priority") == "High")),
        ("Moderate Priority (Score 10-14)", sum(1 for p in properties if p.get("_priority") == "Moderate")),
        ("Watchlist (Score 5-9)", sum(1 for p in properties if p.get("_priority") == "Watchlist")),
        ("", ""),
        ("DISTRESS SIGNALS", ""),
        ("Tax Liens (NYC Finance)", sum(1 for p in properties if p.get("_tax_lien"))),
        ("Lis Pendens / Judgments (any)", sum(1 for p in properties if p.get("_acris_lis_pendens"))),
        ("Lis Pendens < 90 Days (recent)", sum(1 for p in properties if p.get("_acris_lis_pendens_recent"))),
        ("Federal/IRS Liens", sum(1 for p in properties if p.get("_acris_federal_lien"))),
        ("ACRIS Tax Lien Certificates", sum(1 for p in properties if p.get("_acris_tax_lien_cert"))),
        ("Estate/Probate Signals", sum(1 for p in properties if p.get("_acris_estate"))),
        ("HPD Violations 5+", sum(1 for p in properties if (p.get("_hpd_violations") or 0) >= 5)),
        ("HPD Violations 10+", sum(1 for p in properties if (p.get("_hpd_violations") or 0) >= 10)),
        ("Demo Permit on Block", sum(1 for p in properties if p.get("_block_demo"))),
        ("New Building Permit on Block", sum(1 for p in properties if p.get("_block_new_building"))),
        ("LLC Deed Transfer on Block", sum(1 for p in properties if p.get("_block_llc_deed"))),
        ("DOM > 180 Days", sum(1 for p in properties if (p.get("daysOnMarket") or 0) > 180)),
        ("In Block Cluster", sum(1 for p in properties if p.get("_in_cluster"))),
        ("StreetEasy Price Drop >10%", sum(1 for p in properties if (p.get("se_price_drop_pct") or 0) > 10)),
        ("StreetEasy 3+ Cycles", sum(1 for p in properties if (p.get("se_cycles") or 0) >= 3)),
        ("", ""),
        ("NEW SIGNALS", ""),
        ("311 Complaints 10+", sum(1 for p in properties if (p.get("_311_complaints") or 0) >= 10)),
        ("ECB Defaulted Violations", sum(1 for p in properties if (p.get("_ecb_violations") or 0) > 0)),
        ("FDNY Vacate Orders", sum(1 for p in properties if p.get("_fdny_vacate"))),
        ("DOB Complaints 3+", sum(1 for p in properties if (p.get("_dob_complaints") or 0) >= 3)),
        ("New CO on Block", sum(1 for p in properties if p.get("_block_co"))),
        ("CitiBike 5+ Stations", sum(1 for p in properties if (p.get("_citibike_stations") or 0) >= 5)),
    ]

    for row_idx, (label, value) in enumerate(signal_data, 2):
        ws4.cell(row=row_idx, column=1, value=label)
        if label.isupper() or label == "":
            ws4.cell(row=row_idx, column=1).font = Font(bold=True, color="1F4E79")
        ws4.cell(row=row_idx, column=2, value=value if value != "" else "")
        if isinstance(value, int) and value > 0:
            ws4.cell(row=row_idx, column=2).font = Font(bold=True)

    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 15

    # ============================================================
    # Save
    # ============================================================
    output_path = "/tmp/nyc_assemblage/NYC_Assemblage_Scan_2026-03-24.xlsx"
    wb.save(output_path)
    print(f"\n✓ XLSX saved to {output_path}", file=sys.stderr)
    print(f"  Sheets: All Properties ({len(properties)} rows), Immediate Priority, Clusters, Signal Summary", file=sys.stderr)

    print(json.dumps({"path": output_path, "rows": len(properties)}))


if __name__ == "__main__":
    main()
