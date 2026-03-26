#!/usr/bin/env python3
"""
Phase 5: Create Professional XLSX Spreadsheet (Off-Market R8+ Properties)
Color-coded, styled, sorted by composite score.

Input:  /tmp/off_market_scan/verified_properties.json
        /tmp/off_market_scan/clusters.json
Output: /tmp/off_market_scan/Off_Market_R8plus_Scan_{DATE}.xlsx
"""

import json
import sys
import subprocess
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

DATE = datetime.now().strftime("%Y-%m-%d")

# Color schemes
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
IMMEDIATE_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")  # Green
HIGH_FILL = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")       # Light green
MODERATE_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")   # Yellow

BOLD_FONT = Font(bold=True)
LINK_FONT = Font(color="0563C1", underline="single")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF")
)


def fmt_int(val) -> str:
    if val is None or val == "Not stated":
        return "Not stated"
    try:
        v = int(float(str(val).replace(",", "").replace("$", "")))
        return f"{v:,}" if v > 0 else "Not stated"
    except Exception:
        return str(val) if val else "Not stated"


def fmt_bool(val) -> str:
    if val is True:
        return "Yes"
    return "No"


def get_block_context(prop: dict) -> str:
    """Build block context string."""
    parts = []
    if prop.get("_in_cluster"):
        size = prop.get("_cluster_size", "")
        if size:
            parts.append(f"{size} qualifying lots on block")
    if prop.get("_block_llc_deed"):
        parts.append("LLC deed transfer on block (last 12mo)")
    if prop.get("_block_demo"):
        parts.append("Demo permit on block (last 6mo)")
    if prop.get("_block_new_building"):
        parts.append("New building permit on block (last 6mo)")
    return "; ".join(parts) if parts else "None"


def get_notes(prop: dict) -> str:
    """Build notes string."""
    notes = []
    for r in prop.get("_score_reasons", []):
        notes.append(r)
    if prop.get("_data_quality_note"):
        notes.append(f"[!] {prop['_data_quality_note']}")
    return "; ".join(notes) if notes else "Off-market R8+ property"


def main():
    with open("/tmp/off_market_scan/verified_properties.json") as f:
        properties = json.load(f)

    with open("/tmp/off_market_scan/clusters.json") as f:
        clusters = json.load(f)

    print(f"Creating XLSX for {len(properties)} properties...", file=sys.stderr)

    wb = openpyxl.Workbook()

    # ============================================================
    # Sheet 1: All Properties
    # ============================================================
    ws = wb.active
    ws.title = "All Properties"

    headers = [
        "Borough", "Address", "Zoning", "Lot SF", "Bldg SF", "Year Built", "Units",
        "Score", "Priority",
        "Tax Lien?", "Lis Pendens?", "HPD Violations", "Estate?", "Federal Lien?",
        "Demo on Block?", "NB on Block?", "LLC Deed?",
        "311 Complaints", "ECB Violations", "FDNY Vacate?", "DOB Complaints", "Block CO?", "CitiBike",
        "Block Context", "ZoLa Link", "Notes"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    col_widths = {
        1: 12,   # Borough
        2: 40,   # Address
        3: 10,   # Zoning
        4: 10,   # Lot SF
        5: 10,   # Bldg SF
        6: 10,   # Year Built
        7: 8,    # Units
        8: 8,    # Score
        9: 12,   # Priority
        10: 10,  # Tax Lien
        11: 12,  # Lis Pendens
        12: 14,  # HPD Violations
        13: 10,  # Estate
        14: 13,  # Federal Lien
        15: 14,  # Demo on Block
        16: 12,  # NB on Block
        17: 12,  # LLC Deed
        18: 14,  # 311 Complaints
        19: 14,  # ECB Violations
        20: 12,  # FDNY Vacate
        21: 14,  # DOB Complaints
        22: 10,  # Block CO
        23: 10,  # CitiBike
        24: 35,  # Block Context
        25: 15,  # ZoLa
        26: 50,  # Notes
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    for row_idx, prop in enumerate(properties, 2):
        priority = prop.get("_priority", "Watchlist")

        if priority == "Immediate":
            row_fill = IMMEDIATE_FILL
        elif priority == "High":
            row_fill = HIGH_FILL
        elif priority == "Moderate":
            row_fill = MODERATE_FILL
        else:
            row_fill = None

        row_data = [
            prop.get("_borough", "Not stated"),
            prop.get("address", "Not stated"),
            prop.get("_zoning", "Not stated"),
            fmt_int(prop.get("_lot_area")),
            fmt_int(prop.get("_bldg_area")),
            fmt_int(prop.get("_year_built")),
            fmt_int(prop.get("_units_res") or prop.get("_units_total")),
            prop.get("_score", 0),
            priority,
            fmt_bool(prop.get("_tax_lien")),
            fmt_bool(prop.get("_acris_lis_pendens")),
            fmt_int(prop.get("_hpd_violations")),
            fmt_bool(prop.get("_acris_estate")),
            fmt_bool(prop.get("_acris_federal_lien")),
            fmt_bool(prop.get("_block_demo")),
            fmt_bool(prop.get("_block_new_building")),
            fmt_bool(prop.get("_block_llc_deed")),
            fmt_int(prop.get("_311_complaints")),
            fmt_int(prop.get("_ecb_violations")),
            fmt_bool(prop.get("_fdny_vacate")),
            fmt_int(prop.get("_dob_complaints")),
            fmt_bool(prop.get("_block_co")),
            fmt_int(prop.get("_citibike_stations")),
            get_block_context(prop),
            prop.get("_zola_url", "Not stated"),
            get_notes(prop),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = TOP_ALIGN

            if row_fill:
                cell.fill = row_fill

            # Hyperlink for ZoLa
            if col_idx == 25 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.value = "ZoLa"
                cell.font = LINK_FONT

            # Score column: bold + centered
            elif col_idx == 8:
                cell.font = BOLD_FONT
                cell.alignment = CENTER_ALIGN

            # Yes/No columns: center, red bold for Yes
            elif col_idx in [10, 11, 13, 14, 15, 16, 17, 20, 22]:
                cell.alignment = CENTER_ALIGN
                if value == "Yes":
                    cell.font = Font(bold=True, color="C00000")

        ws.row_dimensions[row_idx].height = 20

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ============================================================
    # Sheet 2: Immediate Priority
    # ============================================================
    ws2 = wb.create_sheet("Immediate Priority")
    immediate_props = [p for p in properties if p.get("_priority") == "Immediate"]

    ws2.cell(row=1, column=1, value="IMMEDIATE PRIORITY — NYC Off-Market R8+ Intelligence").font = Font(
        bold=True, size=14, color="1F4E79"
    )
    ws2.cell(row=2, column=1, value=f"Score 15+ | {len(immediate_props)} properties | Generated {DATE}")
    ws2.merge_cells("A1:H1")
    ws2.merge_cells("A2:H2")

    imm_headers = ["Address", "Borough", "Score", "Priority", "Lot SF",
                   "Zoning", "Yr Built", "Tax Lien", "Lis Pendens", "HPD Viol",
                   "Estate", "Block Signal", "Key Signals", "ZoLa"]
    for col_idx, h in enumerate(imm_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, prop in enumerate(immediate_props, 4):
        reasons_short = "; ".join(prop.get("_score_reasons", []))[:120]
        row_data = [
            prop.get("address", ""),
            prop.get("_borough", ""),
            prop.get("_score", 0),
            prop.get("_priority", ""),
            fmt_int(prop.get("_lot_area")),
            prop.get("_zoning", ""),
            fmt_int(prop.get("_year_built")),
            fmt_bool(prop.get("_tax_lien")),
            fmt_bool(prop.get("_acris_lis_pendens")),
            fmt_int(prop.get("_hpd_violations")),
            fmt_bool(prop.get("_acris_estate")),
            get_block_context(prop)[:80],
            reasons_short,
            prop.get("_zola_url", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = IMMEDIATE_FILL
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
            if col_idx == 14 and value and value.startswith("http"):
                cell.hyperlink = value
                cell.value = "ZoLa"
                cell.font = LINK_FONT

    for col, width in enumerate([45, 12, 8, 10, 10, 10, 8, 8, 10, 8, 8, 35, 60, 10], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    # ============================================================
    # Sheet 3: Clusters
    # ============================================================
    ws3 = wb.create_sheet("Block Clusters")
    ws3.cell(row=1, column=1, value="BLOCK CLUSTER OPPORTUNITIES — 2+ Off-Market Properties on Same Block")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E79")
    ws3.merge_cells("A1:G1")

    cluster_headers = ["Borough", "Block", "# Properties", "Addresses",
                       "Total Lot SF", "Zones", "Max Score", "Avg Score"]
    for col_idx, h in enumerate(cluster_headers, 1):
        cell = ws3.cell(row=2, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, cluster in enumerate(clusters, 3):
        total_lot = cluster.get("total_lot_area_sf", "")
        total_lot_str = fmt_int(total_lot) if isinstance(total_lot, (int, float)) else "Unable to verify"

        row_data = [
            cluster.get("borough", ""),
            cluster.get("block", ""),
            cluster.get("property_count", 0),
            "\n".join(cluster.get("addresses", [])),
            total_lot_str,
            ", ".join(cluster.get("zones", [])),
            cluster.get("max_score", 0),
            cluster.get("avg_score", 0),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN

    for col, width in enumerate([12, 12, 10, 60, 14, 20, 10, 10], 1):
        ws3.column_dimensions[get_column_letter(col)].width = width

    # ============================================================
    # Sheet 4: Signal Summary
    # ============================================================
    ws4 = wb.create_sheet("Signal Summary")
    ws4.cell(row=1, column=1, value="OFF-MARKET R8+ DISTRESS SIGNAL SUMMARY").font = Font(
        bold=True, size=14, color="1F4E79"
    )
    ws4.merge_cells("A1:C1")

    signal_data = [
        ("Total R8+ Off-Market Properties Analyzed", len(properties)),
        ("Immediate Priority (Score 15+)", sum(1 for p in properties if p.get("_priority") == "Immediate")),
        ("High Priority (Score 10-14)", sum(1 for p in properties if p.get("_priority") == "High")),
        ("Moderate Priority (Score 6-9)", sum(1 for p in properties if p.get("_priority") == "Moderate")),
        ("Watchlist (Score 1-5)", sum(1 for p in properties if p.get("_priority") == "Watchlist")),
        ("", ""),
        ("DISTRESS SIGNALS", ""),
        ("Tax Liens (NYC Finance)", sum(1 for p in properties if p.get("_tax_lien"))),
        ("Lis Pendens / Judgments (any)", sum(1 for p in properties if p.get("_acris_lis_pendens"))),
        ("Lis Pendens < 90 Days (recent)", sum(1 for p in properties if p.get("_acris_lis_pendens_recent"))),
        ("Federal/IRS Liens", sum(1 for p in properties if p.get("_acris_federal_lien"))),
        ("ACRIS Tax Lien Certificates", sum(1 for p in properties if p.get("_acris_tax_lien_cert"))),
        ("Estate/Probate Signals", sum(1 for p in properties if p.get("_acris_estate"))),
        ("HPD Violations 5+", sum(1 for p in properties if (p.get("_hpd_violations") or 0) >= 5)),
        ("HPD Violations 10+", sum(1 for p in properties if (p.get("_hpd_violations") or 0) >= 10)),
        ("Demo Permit on Block", sum(1 for p in properties if p.get("_block_demo"))),
        ("New Building Permit on Block", sum(1 for p in properties if p.get("_block_new_building"))),
        ("LLC Deed Transfer on Block", sum(1 for p in properties if p.get("_block_llc_deed"))),
        ("In Block Cluster", sum(1 for p in properties if p.get("_in_cluster"))),
        ("", ""),
        ("ADDITIONAL SIGNALS", ""),
        ("311 Complaints 10+", sum(1 for p in properties if (p.get("_311_complaints") or 0) >= 10)),
        ("ECB Defaulted Violations", sum(1 for p in properties if (p.get("_ecb_violations") or 0) > 0)),
        ("FDNY Vacate Orders", sum(1 for p in properties if p.get("_fdny_vacate"))),
        ("DOB Complaints 3+", sum(1 for p in properties if (p.get("_dob_complaints") or 0) >= 3)),
        ("New CO on Block", sum(1 for p in properties if p.get("_block_co"))),
        ("CitiBike 5+ Stations", sum(1 for p in properties if (p.get("_citibike_stations") or 0) >= 5)),
    ]

    for row_idx, (label, value) in enumerate(signal_data, 2):
        ws4.cell(row=row_idx, column=1, value=label)
        if label.isupper() or label == "":
            ws4.cell(row=row_idx, column=1).font = Font(bold=True, color="1F4E79")
        ws4.cell(row=row_idx, column=2, value=value if value != "" else "")
        if isinstance(value, int) and value > 0:
            ws4.cell(row=row_idx, column=2).font = Font(bold=True)

    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 15

    # ============================================================
    # Save
    # ============================================================
    output_path = f"/tmp/off_market_scan/Off_Market_R8plus_Scan_{DATE}.xlsx"
    wb.save(output_path)
    print(f"\nXLSX saved to {output_path}", file=sys.stderr)
    print(f"  Sheets: All Properties ({len(properties)} rows), Immediate Priority, Block Clusters, Signal Summary", file=sys.stderr)

    print(json.dumps({"path": output_path, "rows": len(properties)}))


if __name__ == "__main__":
    main()
